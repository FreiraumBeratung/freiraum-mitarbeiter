import csv, io, re
from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from pydantic import BaseModel
from .db import SessionLocal
from .models import Lead
from .license import has_feature
import openpyxl

router = APIRouter(prefix="/api/leads", tags=["leads"])

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def _valid_email(s: str) -> bool:
    s = (s or "").strip()
    return bool(EMAIL_RE.match(s)) if s else False

def _exists(db, company: str, email: str) -> bool:
    q = db.query(Lead).filter(Lead.company==company.strip())
    if email:
        q = q.filter(Lead.contact_email==(email.strip()))
    return q.first() is not None

def _upsert(db, company: str, name: str, email: str):
    company = (company or "").strip()
    name = (name or "").strip()
    email = (email or "").strip()
    if not company:
        return False
    if email and not _valid_email(email):
        email = ""  # ungültige Mail ignorieren
    if _exists(db, company, email):
        return False
    db.add(Lead(company=company, contact_name=name, contact_email=email, status="new"))
    return True

@router.post("/import/csv")
async def import_csv(file: UploadFile = File(...), 
                     map_company: str = Form("company"),
                     map_contact_name: str = Form("contact_name"),
                     map_contact_email: str = Form("contact_email")):
    if not has_feature("csv_import"):
        raise HTTPException(403, "Feature not allowed in current tier")
    db = SessionLocal()
    try:
        content = await file.read()
        reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
        count = 0
        for row in reader:
            ok = _upsert(db,
                row.get(map_company,""),
                row.get(map_contact_name,""),
                row.get(map_contact_email,"")
            )
            if ok: count += 1
        db.commit()
        return {"ok": True, "imported": count}
    finally:
        db.close()

@router.post("/import/xlsx")
async def import_xlsx(file: UploadFile = File(...),
                      map_company: str = Form("company"),
                      map_contact_name: str = Form("contact_name"),
                      map_contact_email: str = Form("contact_email")):
    if not has_feature("xlsx_import"):
        raise HTTPException(403, "Feature not allowed in current tier (need PRO)")
    db = SessionLocal()
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [ (ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1) ]
        try:
            idx_company = headers.index(map_company)
        except ValueError:
            raise HTTPException(400, "Mapping: map_company header not found")
        idx_name = headers.index(map_contact_name) if map_contact_name in headers else None
        idx_email = headers.index(map_contact_email) if map_contact_email in headers else None
        count = 0
        for r in range(2, ws.max_row+1):
            company = str(ws.cell(row=r, column=(idx_company+1)).value or "").strip()
            name = str(ws.cell(row=r, column=(idx_name+1)).value or "").strip() if idx_name is not None else ""
            email = str(ws.cell(row=r, column=(idx_email+1)).value or "").strip() if idx_email is not None else ""
            ok = _upsert(db, company, name, email)
            if ok: count += 1
        db.commit()
        return {"ok": True, "imported": count}
    finally:
        db.close()

@router.get("/")
def list_leads(limit: int = 50):
    db = SessionLocal()
    try:
        q = db.query(Lead).order_by(Lead.created_at.desc()).limit(limit).all()
        return [ {"id":x.id,"company":x.company,"status":x.status,"email":x.contact_email} for x in q ]
    finally:
        db.close()

class LeadCreate(BaseModel):
    company: str
    contact_name: str = ""
    contact_email: str = ""
    status: str = "new"
    notes: str = ""

@router.post("/")
def create_lead(inp: LeadCreate):
    db = SessionLocal()
    try:
        if inp.contact_email and not _valid_email(inp.contact_email):
            raise HTTPException(400, "Invalid email format")
        if _exists(db, inp.company, inp.contact_email):
            raise HTTPException(409, "Lead already exists")
        lead = Lead(
            company=inp.company.strip(),
            contact_name=inp.contact_name.strip(),
            contact_email=inp.contact_email.strip() if inp.contact_email else None,
            status=inp.status,
            notes=inp.notes
        )
        db.add(lead)
        db.commit()
        return {"ok": True, "id": lead.id}
    finally:
        db.close()
