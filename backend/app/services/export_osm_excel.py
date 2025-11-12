import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_osm_excel(leads: List[Dict], category: str = "", city: str = "") -> str:
    """Exportiert OSM-Leads als Excel mit Raw, Pivot_by_City und Top10_by_Score Sheets"""
    
    # Export-Verzeichnis
    export_dir = Path(os.getenv("EXPORT_DIR", "backend/data/exports"))
    if not export_dir.is_absolute():
        backend_root = Path(__file__).resolve().parents[2]
        export_dir = (backend_root / export_dir).resolve()
    export_dir.mkdir(parents=True, exist_ok=True)
    
    # Dateiname
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"osm_leads_{category}_{city}_{timestamp}.xlsx"
    filepath = export_dir / filename
    
    # Workbook erstellen
    wb = Workbook()
    
    # Sheet 1: Raw
    ws_raw = wb.active
    ws_raw.title = "Raw"
    
    # Headers
    headers = ["Firma", "Kategorie", "Stadt", "Straße", "PLZ", "Telefon", "E-Mail", "Website", "Score", "Quelle", "Lat", "Lon"]
    ws_raw.append(headers)
    
    # Style Headers
    header_fill = PatternFill(start_color="FF7300", end_color="FF7300", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Daten
    for lead in leads:
        ws_raw.append([
            lead.get("company", ""),
            lead.get("category", ""),
            lead.get("city", ""),
            lead.get("street", ""),
            lead.get("postcode", ""),
            lead.get("phone", ""),
            lead.get("email", ""),
            lead.get("website", ""),
            lead.get("score", 0),
            lead.get("source", "osm"),
            lead.get("lat"),
            lead.get("lon"),
        ])
    
    # Auto-Adjust Column Widths
    for col in ws_raw.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[col_letter].width = adjusted_width
    
    # Sheet 2: Pivot_by_City
    ws_pivot = wb.create_sheet("Pivot_by_City")
    
    # Pivot: Group by City
    city_counts: Dict[str, int] = {}
    city_scores: Dict[str, List[int]] = {}
    
    for lead in leads:
        city_name = lead.get("city", "Unbekannt")
        score = lead.get("score", 0)
        
        if city_name not in city_counts:
            city_counts[city_name] = 0
            city_scores[city_name] = []
        
        city_counts[city_name] += 1
        city_scores[city_name].append(score)
    
    # Headers
    ws_pivot.append(["Stadt", "Anzahl Leads", "Ø Score", "Max Score", "Min Score"])
    for col_num in range(1, 6):
        cell = ws_pivot.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Pivot Data
    for city_name in sorted(city_counts.keys()):
        count = city_counts[city_name]
        scores = city_scores[city_name]
        avg_score = sum(scores) / len(scores) if scores else 0
        max_score = max(scores) if scores else 0
        min_score = min(scores) if scores else 0
        
        ws_pivot.append([
            city_name,
            count,
            round(avg_score, 2),
            max_score,
            min_score,
        ])
    
    # Auto-Adjust Column Widths
    for col in ws_pivot.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_pivot.column_dimensions[col_letter].width = adjusted_width
    
    # Sheet 3: Top10_by_Score
    ws_top10 = wb.create_sheet("Top10_by_Score")
    
    # Sort by Score
    sorted_leads = sorted(leads, key=lambda x: x.get("score", 0), reverse=True)
    top10 = sorted_leads[:10]
    
    # Headers
    top10_headers = ["Rank", "Firma", "Stadt", "Telefon", "E-Mail", "Website", "Score"]
    ws_top10.append(top10_headers)
    for col_num in range(1, len(top10_headers) + 1):
        cell = ws_top10.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Top 10 Data
    for rank, lead in enumerate(top10, 1):
        ws_top10.append([
            rank,
            lead.get("company", ""),
            lead.get("city", ""),
            lead.get("phone", ""),
            lead.get("email", ""),
            lead.get("website", ""),
            lead.get("score", 0),
        ])
    
    # Auto-Adjust Column Widths
    for col in ws_top10.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_top10.column_dimensions[col_letter].width = adjusted_width
    
    # Save
    wb.save(filepath)
    
    # Return filename only (for API endpoint)
    return filename

